from django.shortcuts import render,HttpResponse
from django.http import JsonResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from regi_check import serializer,check
from regi_check.config import data_list
from regi_check.models import Phone_Info
import json
from collections import defaultdict
from django.db.models import Q
import openpyxl
from urllib.parse import unquote
import warnings
warnings.filterwarnings('ignore')
reg_dict=[]

# Create your views here.
class checkview(APIView):
    def post(self,request):
        reg_dict.clear()
        # print(request.data)
        phone_nums = request.data.get('phoneNum') 
        post_data=str(request.data)
        print(post_data)
        if len(phone_nums) >=2:
            
            for phone_num in phone_nums:
                
                phone_valid=serializer.validate_phone_number(phone_num)
                if len(phone_num)==11 and phone_valid:
                                    
                    s=check.Spider(phone_num,data_list,post_data)
                    try:    
                        s.run()
                        check.isreg.update({'queryType':1})
                        check.isreg.update({'status':0})
                        reg_dict.append(dict(check.isreg))
                    except:
                        
                        pass
                else:
                    reg_dict.append({
                "phoneNum":phone_num,
                "baidu":'-',
                "aiqiyi":'-',
                "jianshu":'-',
                "weibo":'-',
                'queryType':1,
                "status":1
                
                })
                    
            for item in reg_dict:
                Phone_Info.objects.create(telephone=item.get('phoneNum'),reg_info=json.dumps(item),check_time=request.data.get('create_time'))
                item.update({'create_time':request.data.get('create_time')})  
                
            
            
            
        else:
            phone_valid=serializer.validate_phone_number(phone_nums[0])
            if len(phone_nums[0])==11 and phone_valid:
                s=check.Spider(phone_nums[0],data_list,post_data)
                try:    
                    s.run()
                    check.isreg.update({'queryType':0})
                    if 'status' not in check.isreg:
                        check.isreg.update({'status':0})
                    reg_dict.append(dict(check.isreg))
                except:
                    # print("报错")
                    pass
            else:
                reg_dict.append({
            "phoneNum":phone_nums[0],
            "baidu":'-',
            "aiqiyi":'-',
            "jianshu":'-',
            "weibo":'-',
            'queryType':0,
            "status":1
            
            })
            for item in reg_dict:
                Phone_Info.objects.create(telephone=item.get('phoneNum'),reg_info=str(item),check_time=request.data.get('create_time'))
                item.update({'create_time':request.data.get('create_time')})    
            
        return Response(reg_dict)
    def delete(self,request):
        Phone_Info.objects.all().delete()
        return JsonResponse({'message':'Record deleted successfully!'},status=200)
    

class checkdetailview(APIView):
    def delete(self,request,create_time):
        
        create_time=unquote(create_time)
        
        Phone_Info.objects.filter(check_time=create_time).delete()
        
        return JsonResponse({'message':'Record deleted successfully!'},status=200)
    
    

#仅做测试用    
class checktestview(APIView):
    def get(self,request):
        info_list=Phone_Info.objects.all()
        
        result=serializer.set_to_list(info_list)
        
        
       
        serializer1 = serializer.NestedlistSerializer(instance=result,many=True)
        return Response(serializer1.data)
        # print(serializer1.data)
        
    def post(self,request):
        page = request.data.get('page')
        page_size=request.data.get('page_size')
        phoneNum=request.data.get('phoneNum')
        start_time=request.data.get('start_time')
        end_time=request.data.get('end_time')
       
        skip=(page-1)*page_size
        if phoneNum != None  and start_time != None:
            info_list=Phone_Info.objects.filter(
                Q(telephone__icontains=phoneNum)&Q(check_time__range=[start_time,end_time]))
        elif start_time !=None:
            info_list=Phone_Info.objects.filter(Q(check_time__range=[start_time,end_time]))
        elif phoneNum!=None:
            info_list=Phone_Info.objects.filter(Q(telephone__icontains=phoneNum))
        else:
            info_list=Phone_Info.objects.all()
            
        check_time_set=set(info_list.values_list('check_time',flat=True))
        info_list = Phone_Info.objects.filter(check_time__in=check_time_set)
            
        result=serializer.set_to_list(info_list)
        lenth=len(result)
        total_page=(lenth+page_size-1)//page_size
        paged_info_list = result[skip:skip+page_size]
        serializer1 = serializer.NestedlistSerializer(instance=paged_info_list,many=True)
        modi_data=list(serializer1.data)
        modi_data.insert(0,[{'count':lenth,'pages':total_page}])
        # serializer1.data=modi_data
        # print(serializer1.data)
        return Response(modi_data)

class checkhistoryview(APIView):
    def post(self,request):
        page = request.data.get('page')
        page_size=request.data.get('page_size')
        phoneNum=request.data.get('phoneNum')
        start_time=request.data.get('start_time')
        end_time=request.data.get('end_time')
       
        
        skip=(page-1)*page_size
        if phoneNum !=None  and start_time!=None:
            info_list=Phone_Info.objects.filter(
                Q(telephone__icontains=phoneNum)&Q(check_time__range=[start_time,end_time]))
        elif start_time !=None:
            info_list=Phone_Info.objects.filter(Q(check_time__range=[start_time,end_time]))
        elif phoneNum!=None:
            info_list=Phone_Info.objects.filter(Q(telephone__icontains=phoneNum))
        else:
            info_list=Phone_Info.objects.all()
            
        check_time_set=set(info_list.values_list('check_time',flat=True))
        info_list = Phone_Info.objects.filter(check_time__in=check_time_set)
            
        result=serializer.set_to_list(info_list)
        lenth=len(result)
        total_page=(lenth+page_size-1)//page_size
        paged_info_list = result[skip:skip+page_size]
        serializer1 = serializer.NestedlistSerializer(instance=paged_info_list,many=True)
        modi_data=list(serializer1.data)
        modi_data.insert(0,[{'count':lenth,'pages':total_page}])
        return Response(modi_data)
    
class checkexportview(APIView):
    def post(self,request):
        create_time=request.data.get('create_time')
        info_list=Phone_Info.objects.filter(check_time=create_time)
        result=serializer.set_to_list(info_list)
   
        column_mapping = {'phoneNum': '手机号码',
                          'baidu': '百度',
                          'aiqiyi': '爱奇艺',
                          'jianshu': '简书',
                          'weibo': '微博',
                          
                          'status': '格式(错误为1,频繁为2)',
                          }
       

        
        
        workbook=openpyxl.Workbook()
        worksheet=workbook.active
        column = 1
        for key in result[0][0].reg_info.keys():
            if key !='create_time' and key!='queryType':
                chinese_title=column_mapping.get(key,key)
                worksheet.cell(row=1,column=column,value=chinese_title)
                column+=1
        
        row=2
        for item in result[0]:
            column=1
            for key,value in item.reg_info.items():
                if key !='create_time' and key!='queryType':
                    worksheet.cell(row=row,column=column,value=value)
                    column+=1
            row+=1 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="records.xlsx"' 
        workbook.save(response)

        return response